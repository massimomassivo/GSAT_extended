# Import external dependencies
import sys, os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NumberFormatDescriptor
from skimage import morphology as morph
from skimage.util import img_as_ubyte, img_as_bool
from skimage.util import invert as ski_invert
from skimage import transform as tran
from matplotlib import pyplot as plt

# Import local modules
# Ensure this is the correct path to the functions folder
sys.path.insert(1, '../imppy3d_functions')
import grain_size_functions as gsz
import import_export as imex
import volume_image_processing as vol
import ski_driver_functions as sdrv

# Set constants related to plotting (for MatPlotLib)
SMALL_SIZE = 10
MEDIUM_SIZE = 12
BIGGER_SIZE = 14

plt.rc('font', size=MEDIUM_SIZE)         # Controls default text sizes
plt.rc('axes', titlesize=BIGGER_SIZE)    # Fontsize of the axes title
plt.rc('axes', labelsize=MEDIUM_SIZE)    # Fontsize of the x and y labels
plt.rc('xtick', labelsize=MEDIUM_SIZE)   # Fontsize of the tick labels
plt.rc('ytick', labelsize=MEDIUM_SIZE)   # Fontsize of the tick labels
plt.rc('legend', fontsize=MEDIUM_SIZE)   # Legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # Fontsize of the figure title


# -------- USER INPUTS --------

# Provide the filepath to the image that should be imported. This image should
# be a binary image (containing only black and white pixels), and it should be
# saved as an unsigned 8-bit image. Moreover, the metallographic grain
# boundaries are assumed to be denoted by white pixels.
file_in_path = "path/to/segmented_image.jpg"

# Define the path for the second Excel file to compare results between images
second_excel_path = "path/to/summary.xlsx"

# Lines of pixels will be drawn onto rotated versions of the images. The 
# distances between intersections, in pixels, correspond to grain lengths.
# The distances of these individual segments as a function of rotation angle
# will be saved in an automatically named .csv file. It will be saved in the
# same directory as the imported image, and it will be appended with 
# "_distances". For example, if the imported image was, "./img_segmented.tif",
# then the output file would be saved as, "./img_segmented_distances.csv".

# Provide a path to a directory (not a file) where images can be saved. These
# images will be rotated copies of the provided input image, but with
# superimposed lines on them that illustrate the intersect-segments. There
# will be one image for every rotation, and the images will be saved as unsigned
# integer (8-bit) images. Like the input image, grain boundaries are denoted by
# white pixels. The intersect segments are superimposed as gray pixels with
# intensity 150 (out of 255). The names of these saved images are automatically
# generated based on the name of the input image and the corresponding rotation
# angles.
directory_images_out = "path/to/output_images"

# Set "borders_white" to True if the provided input image contains white pixels
# that correspond to the grain boundaries. If True, then nothing is done. If
# False, then it is assumed that the base color of the image is white with grain
# boundaries denoted by black pixels. In this case, when False, the image will
# be inverted such that grain boundaries are denoted by white pixels.
borders_white = True

# The distance, in pixels, between each subsequent line to be drawn is given by
# "row_step", which should be a positive integer greater than zero. A smaller
# "row_step" corresponds to more lines being drawn leading to more intersection
# segments.
row_step = 20 # pixels

# The provided binary image will be rotated to check the length of grains along
# different directions. The rotations that will be performed depend on the input
# parameters "theta_start", "theta_end", "n_theta_steps", and 
# "inclusive_theta_end". These parameters are wrappers for numpy.linspace().
# Set "theta_start" and "theta_end" as the starting and ending angles, 
# specifically as positive floats. If "inclusive_theta_end" is False, then 
# the end point "theta_end" will not be included in the sequence of rotations. 
# "n_theta_steps" defines the number of theta rotations to generate between
# "theta_start" and "theta_end", and as before, "n_theta_steps" should be a
# positive float.
theta_start = 0.0 # degrees
theta_end = 180 # degrees
n_theta_steps = 6
inclusive_theta_end = False

# If the provided binary image has been skeletonized, then after each image
# rotation that is applied in this algorithm below, it is advised that the
# image undergoes a small morphological dilation and gets skeletonized again
# to ensure adequate continuity. Set "reskeletonize" to True in order to perform
# a small dilation followed by a skeletonization after each image rotation.
reskeletonize = True

# Define how many pixels correspond to a certain amount of µm (micrometers)
scalebar_pixel = 464
scalebar_micrometer = 50
pixels_per_micron = scalebar_pixel/scalebar_micrometer  # Example: 10 pixels correspond to 1 µm

# Define the start and end pixels for height and width to cut off scalebar
start_row, end_row = 0, 1825
start_col, end_col = 0, 2580

# -------- IMPORT IMAGE FILE --------

img1, img1_prop = imex.load_image(file_in_path)

if img1 is None:
    print(f"\nFailed to import images from the directory: \n{file_in_path}")
    print("\nQuitting the script...")
    quit()

# Optionally, extract the (Numpy) properties of image.
img1 = img_as_ubyte(img1)
img1_size = img1_prop[0]  # Total number of pixels
img1_shape = img1_prop[1] # Tuple containing the number of rows and columns
img1_dtype = img1_prop[2] # Returns the image data type (i.e., uint8)

img2temp = img1.copy()
img2=img2temp[start_row:end_row, start_col:end_col]

n_row = img2.shape[0]
n_col = img2.shape[1]


# -------- CALCULATE DISTANCE BETWEEN BINARY PIXELS --------

if not borders_white:
    img2 = ski_invert(img2)

row_step = int(np.absolute(row_step))
if row_step > n_row:
    row_step = n_row - 1
elif row_step == 0:
    row_step = 1

# Padd the image to account for growing size during rotations
circle_min_diameter = int(np.ceil(np.sqrt(n_row**2 + n_col**2)))
n_pad_temp = 1.1*(circle_min_diameter - np.amin(img2.shape))*0.5
n_pad = int(np.ceil(n_pad_temp))
img2 = vol.pad_image_boundary(img2, cval_in=0, n_pad_in=n_pad)
n_row2 = img2.shape[0]
n_col2 = img2.shape[1]

# All pixel distances will be stored in dist_all_arr along with angles
# The shape will be [n,2] where n is the number of segment lengths
# that were found. The first column will be the image rotation that
# was used for the corresponding segment length in the second column.
dist_all_arr = []
inv_dist_all_arr = []

# Each rotated image with the superimposed intersect lines will be
# saved in the following list.
imgs_rot_out = []

print(f"\nCounting intersect distances for each orientation...")
n_theta_steps = int(np.round(n_theta_steps))
theta_arr = np.linspace(theta_start, theta_end, n_theta_steps, \
                        endpoint=inclusive_theta_end)
img_rot = img2.copy()
for cur_theta in theta_arr:

    if cur_theta != 0:
        img_rot = img_as_ubyte(tran.rotate(img2, cur_theta, order=1, resize=False))
        img_rot[img_rot >= 128] = 255
        img_rot[img_rot < 128] = 0

        if reskeletonize:
            op_type = 2   # binary_closing
            foot_type = 1 # disk kernel
            morph_rad = 1 # radius 2 pixels
            filt_params = [int(op_type), int(foot_type), int(morph_rad)]
            img_rot = sdrv.apply_driver_morph(img_rot, filt_params, quiet_in=True)

            img_bool = img_as_bool(img_rot)
            img_rot = morph.skeletonize(img_bool)
            img_rot = img_as_ubyte(img_rot)

    # Select various rows of interest and count pixels between intersects
    for rr in np.arange(0, n_row2, row_step):

        # Get the start and end indices of the grain segments
        cur_row = img_rot[rr]
        segment_ii = gsz.find_intersections(cur_row)

        # Loop through segments and calculate the distances and color the image
        for cur_seg in segment_ii:
            pix_dist = cur_seg[1] - cur_seg[0]
            micron_dist = pix_dist / pixels_per_micron  # Convert pixel distance to microns
            dist_all_arr.append([cur_theta, micron_dist])
            inv_dist_all_arr.append([cur_theta, 1 / micron_dist])  # Calculate and store inverse distance in microns
            img_rot[rr, cur_seg[0]:cur_seg[1]] = 150

    imgs_rot_out.append(img_rot)

    print(f"  Completed processing intersects for the {cur_theta:.2f} deg orientation...")

dist_all_arr = np.array(dist_all_arr)
inv_dist_all_arr = np.array(inv_dist_all_arr)



# -------- SAVE ROTATED IMAGES --------

# for m, cur_theta in enumerate(theta_arr):
#     cur_img = imgs_rot_out[m]
#
#     # Extract the filename without the path
#     filename_no_ext = os.path.basename(os.path.splitext(file_in_path)[0])
#     file_out_name = filename_no_ext + "_" + f"{int(np.round(cur_theta))}" + ".tif"
#     cur_img_path_out = os.path.join(directory_images_out, file_out_name)
#
#     save_flag = imex.save_image(cur_img, cur_img_path_out)
#     if not save_flag:
#         print(f"\nCould not save {cur_img_path_out}")
#
# print(f"Rotated images have been successfully saved to {directory_images_out}.")
    

# -------- SAVE DATA TO CSV FILE -------- (auskommentiert, habe meine eigene Excel erstellt)

# filename_no_ext, file_extension = os.path.splitext(file_in_path)
# csv_file_out_name = filename_no_ext + "_invdistances.csv"
#
# header_str = ["Angle (deg)", "Segment Distance (Pixels)"]
# gsz.save_csv(inv_dist_all_arr, csv_file_out_name, header_str)


# -------- PLOT THE DATA FOR EACH ORIENTATION --------

print(f"\nCreating a box-plot associated with each rotation...")

# Reshape the data to make it suitable for plotting
theta_diff_arr = np.diff(dist_all_arr[:,0])
theta_step_max = np.amax(theta_diff_arr)

n_segs = dist_all_arr.shape[0]
theta_2d_plt = []
pix_dist_2d_plt = []
inv_pix_dist_2d_plt = []  # Array to store inverse distances
arr_start_i = 0
for m in np.arange(n_segs-1):
    theta_m = dist_all_arr[m,0]
    theta_mpo = dist_all_arr[m+1,0]

    if np.absolute(theta_mpo - theta_m) >= theta_step_max/2.0:
        temp_arr = dist_all_arr[arr_start_i:m+1,:]
        temp_arr_inv = inv_dist_all_arr[arr_start_i:m+1,:]
        theta_2d_plt.append(temp_arr[:,0])
        pix_dist_2d_plt.append(temp_arr[:,1])
        inv_pix_dist_2d_plt.append(temp_arr_inv[:, 1])  # Calculate inverse distances
        arr_start_i = m + 1

temp_arr = dist_all_arr[arr_start_i:n_segs,:]
temp_arr_inv = inv_dist_all_arr[arr_start_i:n_segs,:]
theta_2d_plt.append(temp_arr[:,0])
pix_dist_2d_plt.append(temp_arr[:,1])
inv_pix_dist_2d_plt.append(temp_arr_inv[:, 1])  # Calculate inverse distances

theta_labels_1D = []
for cur_theta_arr in theta_2d_plt:
    temp_str = f"{cur_theta_arr[0]:.1f}"
    theta_labels_1D.append(temp_str)

fig1, ax1 = plt.subplots()
ax1.boxplot(pix_dist_2d_plt, sym="", labels=theta_labels_1D)

num_boxes = len(pix_dist_2d_plt)
for m in range(num_boxes):
    cur_y_arr = pix_dist_2d_plt[m]
    cur_x_arr = np.random.normal(m+1, 0.05, size=cur_y_arr.size)
    ax1.plot(cur_x_arr, cur_y_arr, '.b', alpha=0.2)

ax1.set_xlabel('Rotation of Line Intercepts [degrees]')
ax1.set_ylabel('Segment Length of Grains [µm]')  # pixels to µm

# Define the input file path and create the output boxplot file path
input_file_path_no_ext, _ = os.path.splitext(file_in_path)
output_boxplot_file = input_file_path_no_ext + "_boxplot.png"

# Speichere den Boxplot
fig1.savefig(output_boxplot_file)

print(f"\nBoxplot has been successfully saved to {output_boxplot_file}.")

# -------- SUMMARIZE THE RESULTS IN THE COMMAND WINDOW AND SAVE TO EXCEL --------

print("\n\n ========== RESULTS SUMMARY ========== ")

results = {
    "Property": [
        "Total Number of Grain Segments",
        "Summed Length of Grain Segments (µm)",
        "Average Grain Size (µm)",
        "Median Grain Size (µm)",
        "Std. Deviation in Grain Size (µm)",
        "Thickness (Average inverse Grain Size) (µm)",
        "Thickness (Median inverse Grain Size) (µm)"
    ]
}

# Define input parameters
input_parameters = {
    "Filename": os.path.basename(file_in_path),
    "borders_white": borders_white,
    "row_step": row_step,
    "theta_start": theta_start,
    "theta_end": theta_end,
    "n_theta_steps": n_theta_steps,
    "inclusive_theta_end": inclusive_theta_end,
    "reskeletonize": reskeletonize,
    "scalebar_pixel": scalebar_pixel,
    "scalebar_micrometer": scalebar_micrometer
}

for m, cur_theta_arr in enumerate(theta_2d_plt):
    cur_dist_arr = pix_dist_2d_plt[m]
    inv_cur_dist_arr = inv_pix_dist_2d_plt[m]  # Inverse distances array
    cur_theta = theta_labels_1D[m]
    temp_sum = np.sum(cur_dist_arr)
    temp_avg = np.mean(cur_dist_arr)
    temp_med = np.median(cur_dist_arr)
    temp_std = np.std(cur_dist_arr)
    inv_temp_avg = np.mean(inv_cur_dist_arr)  # Calculate mean of inverse distances
    inv_temp_med = np.median(inv_cur_dist_arr)  # Calculate median of inverse distances
    thickness_avg = 1 / (1.5 * inv_temp_avg)  # Apply thickness formula to mean inverse distance
    thickness_med = 1 / (1.5 * inv_temp_med)  # Apply thickness formula to median inverse distance

    print(f"\n --- Grain Size Statistics for {cur_theta} deg --- ")
    print(f"  Total Number of Grain Segments: {cur_dist_arr.size}")
    print(f"  Summed Length of Grain Segments (Pixels): {temp_sum:.2f}")
    print(f"  Average Grain Size (Pixels): {temp_avg:.2f}")
    print(f"  Median Grain Size (Pixels): {temp_med:.2f}")
    print(f"  Std. Deviation in Grain Size (Pixels): {temp_std:.2f}")
    #print(f"  Average Inverse Grain Size (Pixels): {inv_temp_avg:.6f}")  # Print average inverse grain size
    #print(f"  Median Inverse Grain Size (Pixels): {inv_temp_med:.6f}")   # Print median inverse grain size
    #print(f"  Thickness from Average Inverse Grain Size (Pixels): {thickness_avg:.2f}")  # Print thickness from average inverse distance
    #print(f"  Thickness from Median Inverse Grain Size (Pixels): {thickness_med:.2f}")  # Print thickness from median inverse distance

    results[cur_theta] = [
        cur_dist_arr.size,
        temp_sum,
        temp_avg,
        temp_med,
        temp_std,
        #inv_temp_avg,  # Add average inverse grain size to results
        #inv_temp_med   # Add median inverse grain size to results
        thickness_avg,  # Add thickness from average inverse distance to results
        thickness_med   # Add thickness from median inverse distance to results
    ]

temp_sum = np.sum(dist_all_arr[:,1])
temp_avg = np.mean(dist_all_arr[:,1])
temp_med = np.median(dist_all_arr[:,1])
temp_std = np.std(dist_all_arr[:,1])
inv_temp_avg_all_lines = np.mean(inv_dist_all_arr[:,1])  # Calculate mean of all inverse distances
inv_temp_med_all_lines = np.median(inv_dist_all_arr[:,1])  # Calculate median of all inverse distances
thickness_avg_all_lines = 1 / (1.5 * inv_temp_avg_all_lines)  # Apply thickness formula to mean inverse distance for all lines
thickness_med_all_lines = 1 / (1.5 * inv_temp_med_all_lines)  # Apply thickness formula to median inverse distance for all lines

print(f"\n --- All Lines Grain Size Statistics --- ")
print(f"  Total Number of Grain Segments: {dist_all_arr.shape[0]}")
print(f"  Summed Length of Grain Segments (µm): {temp_sum:.2f}")
print(f"  Average Grain Size (µm): {temp_avg:.2f}")
print(f"  Median Grain Size (µm): {temp_med:.2f}")
print(f"  Std. Deviation in Grain Size (µm): {temp_std:.2f}")
# print(f"  Average Inverse Grain Size (µm): {inv_temp_avg_all_lines:.6f}")  # Print average inverse grain size for all lines
# print(f"  Median Inverse Grain Size (µm): {inv_temp_med_all_lines:.6f}")   # Print median inverse grain size for all lines
print(f"  Thickness from Average Inverse Grain Size (µm): {thickness_avg_all_lines:.2f}")  # Print thickness from average inverse distance for all lines
print(f"  Thickness from Median Inverse Grain Size (µm): {thickness_med_all_lines:.2f}")   # Print thickness from median inverse distance for all lines

results["All Lines"] = [
    dist_all_arr.shape[0],
    temp_sum,
    temp_avg,
    temp_med,
    temp_std,
    # inv_temp_avg_all_lines,  # Add average inverse grain size for all lines to results
    # inv_temp_med_all_lines   # Add median inverse grain size for all lines to results
    thickness_avg_all_lines,  # Add thickness from average inverse distance for all lines to results
    thickness_med_all_lines  # Add thickness from median inverse distance for all lines to results
]

# Convert the dictionary to a DataFrame
df_results = pd.DataFrame(results)

# Convert dist_all_arr[:,1] and inv_dist_all_arr[:,1] to DataFrames
df_distances = pd.DataFrame(dist_all_arr[:,1], columns=['Distances (µm)'])
df_inv_distances = pd.DataFrame(inv_dist_all_arr[:,1], columns=['Inverse Distances (1/µm)'])

# Define the input file path and create the output Excel file path
input_file_path_no_ext, _ = os.path.splitext(file_in_path)
output_file_path_excel = input_file_path_no_ext + ".xlsx"

# # Save the DataFrame to an Excel file
# df_results.to_excel(output_file_path_excel, index=False, engine='openpyxl')

# Save the DataFrames to separate sheets in the Excel file
with pd.ExcelWriter(output_file_path_excel, engine='openpyxl') as writer:
    df_results.to_excel(writer, sheet_name='Results', index=False)
    df_distances.to_excel(writer, sheet_name='Distances', index=False)
    df_inv_distances.to_excel(writer, sheet_name='Inverse Distances', index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(output_file_path_excel)
ws = wb.active

# Apply number format to the relevant columns
for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
    for cell in row:
        cell.number_format = '0.00'

# Add input parameters below the results table
start_row = ws.max_row + 2  # Leave one empty row
ws.cell(row=start_row, column=1, value="Input Parameters")
for i, (param, value) in enumerate(input_parameters.items(), start=start_row + 1):
    ws.cell(row=i, column=1, value=param)
    ws.cell(row=i, column=2, value=value)

# Save the workbook
wb.save(output_file_path_excel)

print(f"\nResults and input parameters have been successfully saved to {output_file_path_excel}.")

# Define the input parameters and results
summary_parameters = {
    "Filename": os.path.basename(file_in_path),
    "Avg Grain Size (µm)": temp_avg,
    "Med. Grain Size (µm)": temp_med,
    "Thickness (Avg inv. gs) (µm)": thickness_avg_all_lines,
    "Thickness (Med. inv. gs) (µm)": thickness_med_all_lines,
    "Inverse Int Avg": inv_temp_avg_all_lines,
    "Inverse Int Med": inv_temp_med_all_lines
}

# Convert the input parameters to a DataFrame
df_summary = pd.DataFrame([summary_parameters])

# Check if the second Excel file exists
if not os.path.exists(second_excel_path):
    # If the file does not exist, create it and write the DataFrame to it
    df_summary.to_excel(second_excel_path, index=False, engine='openpyxl')
else:
    # If the file exists, append the new data without overwriting the old data
    with pd.ExcelWriter(second_excel_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        df_summary.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

# Load the workbook and select the active worksheet
wb = load_workbook(second_excel_path)
ws = wb.active

# Apply number format to the relevant columns
for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
    for cell in row:
        cell.number_format = '0.00'

# Save the workbook
wb.save(second_excel_path)

print(f"\nSummary data has been successfully saved to {second_excel_path}.")

# Funktion zum Erstellen und Speichern von Histogrammen
def create_histogram(data, title, xlabel, output_file):
    plt.figure()
    plt.hist(data, bins=30, edgecolor='black')
    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel('Frequency')
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(output_file)
    plt.close()

# Definiere den Eingabepfad und erstelle die Ausgabepfade für die Histogramme
input_file_path_no_ext, _ = os.path.splitext(file_in_path)
output_histogram_distances = input_file_path_no_ext + "_histogram_distances.png"
output_histogram_inv_distances = input_file_path_no_ext + "_histogram_inverse_distances.png"

# Erstelle und speichere die Histogramme
create_histogram(dist_all_arr[:,1], 'Histogram of Distances', 'Distance (µm)', output_histogram_distances)
create_histogram(inv_dist_all_arr[:,1], 'Histogram of Inverse Distances', 'Inverse Distance (1/µm)', output_histogram_inv_distances)

print(f"\nHistograms have been successfully saved to {output_histogram_distances} and {output_histogram_inv_distances}.")

#Actually show the plot after the summary has been reported
plt.show()
print("\n\nScript successfully terminated!")
