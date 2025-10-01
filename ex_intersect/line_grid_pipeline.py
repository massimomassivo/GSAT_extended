"""Processing pipeline for line grid intersection measurements."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from skimage import morphology as morph
from skimage import transform as tran
from skimage.util import img_as_bool, img_as_ubyte
from skimage.util import invert as ski_invert

from imppy3d_functions import grain_size_functions as gsz
from imppy3d_functions import import_export as imex
from imppy3d_functions import ski_driver_functions as sdrv
from imppy3d_functions import volume_image_processing as vol


@dataclass
class LineGridConfig:
    """Collect configuration parameters for the line-grid pipeline."""

    file_in_path: Path
    results_base_dir: Path
    summary_excel_path: Optional[Path] = None
    borders_white: bool = True
    row_step: int = 20
    theta_start: float = 0.0
    theta_end: float = 180.0
    n_theta_steps: int = 6
    inclusive_theta_end: bool = False
    reskeletonize: bool = True
    scalebar_pixel: float = 464.0
    scalebar_micrometer: float = 50.0
    crop_rows: Tuple[int, int] = (0, 1825)
    crop_cols: Tuple[int, int] = (0, 2580)

    def __post_init__(self) -> None:
        self.file_in_path = Path(self.file_in_path)
        self.results_base_dir = Path(self.results_base_dir)
        if self.summary_excel_path is None:
            self.summary_excel_path = self.results_base_dir / "summary.xlsx"
        else:
            self.summary_excel_path = Path(self.summary_excel_path)

    @property
    def pixels_per_micron(self) -> float:
        """Return the scale factor that converts pixels to microns."""

        return self.scalebar_pixel / self.scalebar_micrometer

    def to_input_parameters(self) -> Dict[str, object]:
        """Return a serialisable mapping of configuration values."""

        return {
            "Filename": self.file_in_path.name,
            "borders_white": self.borders_white,
            "row_step": self.row_step,
            "theta_start": self.theta_start,
            "theta_end": self.theta_end,
            "n_theta_steps": self.n_theta_steps,
            "inclusive_theta_end": self.inclusive_theta_end,
            "reskeletonize": self.reskeletonize,
            "scalebar_pixel": self.scalebar_pixel,
            "scalebar_micrometer": self.scalebar_micrometer,
            "crop_rows": self.crop_rows,
            "crop_cols": self.crop_cols,
        }


@dataclass
class ImageMetadata:
    """Store metadata of the imported image."""

    size: int
    shape: Tuple[int, int]
    dtype: str


@dataclass
class PreparedImageData:
    """Hold the pre-processed image data used for the measurements."""

    original_image: np.ndarray
    cropped_image: np.ndarray
    padded_image: np.ndarray
    metadata: ImageMetadata
    theta_values: np.ndarray
    results_dir: Path
    rotated_dir: Path
    base_output_path: Path
    row_step: int
    pixels_per_micron: float


@dataclass
class LineMeasurementResult:
    """Container with the measured distances for each rotation angle."""

    distances: np.ndarray
    inverse_distances: np.ndarray
    per_theta_distances: List[np.ndarray]
    per_theta_inverse_distances: List[np.ndarray]
    theta_labels: List[str]
    rotated_images: List[np.ndarray]


@dataclass
class AngleStatistics:
    """Computed statistics for one rotation angle (or all angles combined)."""

    angle_label: str
    segment_count: int
    total_length: float
    average_length: float
    median_length: float
    std_dev: float
    average_inverse: float
    median_inverse: float
    thickness_from_average: float
    thickness_from_median: float
    astm_g: float
    ci95_halfwidth_um: float
    ci95_low_um: float
    ci95_high_um: float
    rel_accuracy_pct: float


def astm_g_from_lbar_um(lbar_um: float) -> float:
    import numpy as np

    if not np.isfinite(lbar_um) or lbar_um <= 0:
        return float("nan")
    l_mm = lbar_um / 1000.0
    g = (-6.643856 * np.log10(l_mm)) - 3.288
    return float(np.round(g, 1))


T95_TABLE = {
    1: 12.706,
    2: 4.303,
    3: 3.182,
    4: 2.776,
    5: 2.571,
    6: 2.447,
    7: 2.365,
    8: 2.306,
    9: 2.262,
    10: 2.228,
    12: 2.179,
    15: 2.131,
    20: 2.086,
    25: 2.060,
    30: 2.042,
    40: 2.021,
    60: 2.000,
    120: 1.980,
    240: 1.970,
    10_000_000: 1.960,
}


def t95_from_df(df: int) -> float:
    if df <= 1:
        return T95_TABLE[1]
    for key in sorted(T95_TABLE.keys()):
        if df <= key:
            return T95_TABLE[key]
    return 1.960


@dataclass
class StatisticsResult:
    """Aggregate results from the measured line intersections."""

    angle_statistics: List[AngleStatistics]
    overall_statistics: AngleStatistics
    results_table: pd.DataFrame
    distances_df: pd.DataFrame
    inverse_distances_df: pd.DataFrame
    summary_row: pd.DataFrame


@dataclass
class SaveOptions:
    """Configure which artefacts should be written to disk."""

    save_rotated_images: bool = False
    save_boxplot: bool = True
    save_histograms: bool = True
    save_excel: bool = True
    append_summary: bool = True
    show_plots: bool = False


@dataclass
class SaveArtifacts:
    """Track the artefacts created by :func:`save_outputs`."""

    excel_path: Optional[Path] = None
    summary_excel_path: Optional[Path] = None
    boxplot_path: Optional[Path] = None
    histogram_paths: List[Path] = field(default_factory=list)
    rotated_image_paths: List[Path] = field(default_factory=list)


def describe_measurements(measurements: "LineMeasurementResult") -> None:
    """Log how many segments were detected for each orientation.

    Parameters
    ----------
    measurements : LineMeasurementResult
        Measured distances and labels returned by :func:`measure_line_intersections`.

    Examples
    --------
    >>> describe_measurements(LineMeasurementResult(  # doctest: +SKIP
    ...     distances=np.empty((0, 2)),
    ...     inverse_distances=np.empty((0, 2)),
    ...     per_theta_distances=[],
    ...     per_theta_inverse_distances=[],
    ...     theta_labels=[],
    ...     rotated_images=[],
    ... ))
    """

    for label, distances in zip(
        measurements.theta_labels, measurements.per_theta_distances
    ):
        count = int(distances.size)
        print(
            "  Completed processing intersects for the "
            f"{float(label):.2f} deg orientation with {count} segments..."
        )


def _print_angle_statistics(stat: "AngleStatistics") -> None:
    """Print statistics for a single rotation angle.

    Parameters
    ----------
    stat : AngleStatistics
        Statistics calculated by :func:`aggregate_statistics`.
    """

    if stat.angle_label == "All Lines":
        header = "\n --- All Lines Grain Size Statistics --- "
    else:
        header = f"\n --- Grain Size Statistics for {stat.angle_label} deg --- "
    print(header)
    print(f"  Total Number of Grain Segments: {stat.segment_count}")
    print(f"  Summed Length of Grain Segments (µm): {stat.total_length:.2f}")
    print(f"  Average Grain Size (µm): {stat.average_length:.2f}")
    print(f"  Median Grain Size (µm): {stat.median_length:.2f}")
    print(f"  Std. Deviation in Grain Size (µm): {stat.std_dev:.2f}")
    print(
        f"  Thickness from Average Inverse Grain Size (µm): {stat.thickness_from_average:.2f}"
    )
    print(
        f"  Thickness from Median Inverse Grain Size (µm): {stat.thickness_from_median:.2f}"
    )
    print(f"  ASTM G (from l̄): {stat.astm_g:.1f}")
    print(
        "  95% CI (±): "
        f"{stat.ci95_halfwidth_um:.3f} µm  -> [{stat.ci95_low_um:.3f}, {stat.ci95_high_um:.3f}] µm"
    )
    print(f"  Relative Accuracy (95%): {stat.rel_accuracy_pct:.2f} %")


def print_statistics(statistics: "StatisticsResult") -> None:
    """Emit a human-readable summary of the computed statistics.

    Parameters
    ----------
    statistics : StatisticsResult
        Aggregated statistics returned by :func:`aggregate_statistics`.

    Examples
    --------
    >>> print_statistics(StatisticsResult(  # doctest: +SKIP
    ...     angle_statistics=[],
    ...     overall_statistics=AngleStatistics(
    ...         angle_label="All Lines",
    ...         segment_count=0,
    ...         total_length=0.0,
    ...         average_length=float("nan"),
    ...         median_length=float("nan"),
    ...         std_dev=float("nan"),
    ...         average_inverse=float("nan"),
    ...         median_inverse=float("nan"),
    ...         thickness_from_average=float("nan"),
    ...         thickness_from_median=float("nan"),
    ...         astm_g=float("nan"),
    ...         ci95_halfwidth_um=float("nan"),
    ...         ci95_low_um=float("nan"),
    ...         ci95_high_um=float("nan"),
    ...         rel_accuracy_pct=float("nan"),
    ...     ),
    ...     results_table=pd.DataFrame(),
    ...     distances_df=pd.DataFrame(),
    ...     inverse_distances_df=pd.DataFrame(),
    ...     summary_row=pd.DataFrame(),
    ... ))
    """

    print("\n\n ========== RESULTS SUMMARY ========== ")
    for angle_stat in statistics.angle_statistics:
        _print_angle_statistics(angle_stat)
    _print_angle_statistics(statistics.overall_statistics)


def prepare_image(config: LineGridConfig) -> PreparedImageData:
    """Load the input image and apply pre-processing steps.

    Parameters
    ----------
    config : LineGridConfig
        Configuration describing the file paths and cropping parameters.

    Returns
    -------
    PreparedImageData
        Pre-processed images, metadata and derived directories used downstream.

    Raises
    ------
    FileNotFoundError
        If the input image cannot be loaded.

    Examples
    --------
    >>> cfg = LineGridConfig(file_in_path="image.png", results_base_dir=Path("./results"))  # doctest: +SKIP
    >>> prepared = prepare_image(cfg)  # doctest: +SKIP
    """

    img1, img1_prop = imex.load_image(str(config.file_in_path))
    if img1 is None:
        raise FileNotFoundError(
            f"Could not load image at '{config.file_in_path}'."
        )

    img1 = img_as_ubyte(img1)
    metadata = ImageMetadata(
        size=int(img1_prop[0]),
        shape=tuple(img1_prop[1]),
        dtype=str(img1_prop[2]),
    )

    start_row, end_row = config.crop_rows
    start_col, end_col = config.crop_cols
    cropped = img1.copy()[start_row:end_row, start_col:end_col]

    if not config.borders_white:
        cropped = ski_invert(cropped)

    row_step = int(abs(config.row_step))
    if row_step == 0:
        row_step = 1
    if row_step >= cropped.shape[0]:
        row_step = max(1, cropped.shape[0] - 1)

    circle_min_diameter = int(np.ceil(np.sqrt(cropped.shape[0] ** 2 + cropped.shape[1] ** 2)))
    n_pad_temp = 1.1 * (circle_min_diameter - np.amin(cropped.shape)) * 0.5
    n_pad = int(np.ceil(n_pad_temp))
    padded = vol.pad_image_boundary(cropped, cval_in=0, n_pad_in=n_pad)

    theta_values = np.linspace(
        config.theta_start,
        config.theta_end,
        int(np.round(config.n_theta_steps)),
        endpoint=config.inclusive_theta_end,
    )

    file_stem = config.file_in_path.stem
    suffix = config.file_in_path.suffix.lower().lstrip(".")
    suffix_part = f"_{suffix}" if suffix else ""
    results_dir = config.results_base_dir / f"{file_stem}{suffix_part}_results"
    rotated_dir = results_dir / "rotated_images"
    base_output_path = results_dir / file_stem

    return PreparedImageData(
        original_image=img1,
        cropped_image=cropped,
        padded_image=padded,
        metadata=metadata,
        theta_values=theta_values,
        results_dir=results_dir,
        rotated_dir=rotated_dir,
        base_output_path=base_output_path,
        row_step=row_step,
        pixels_per_micron=config.pixels_per_micron,
    )


def measure_line_intersections(
    prepared: PreparedImageData, config: LineGridConfig
) -> LineMeasurementResult:
    """Measure all line intersections for the configured rotations.

    Parameters
    ----------
    prepared : PreparedImageData
        Image data returned by :func:`prepare_image`.
    config : LineGridConfig
        Pipeline configuration providing sampling and morphology options.

    Returns
    -------
    LineMeasurementResult
        Distances and intermediate artefacts for each rotation angle.

    Raises
    ------
    ValueError
        Propagated from driver functions when invalid parameters are provided.

    Examples
    --------
    >>> measure_line_intersections(prepared, config)  # doctest: +SKIP
    LineMeasurementResult(...)
    """

    distances: List[Tuple[float, float]] = []
    inverse_distances: List[Tuple[float, float]] = []
    per_theta_distances: List[np.ndarray] = []
    per_theta_inverse: List[np.ndarray] = []
    rotated_images: List[np.ndarray] = []

    for cur_theta in prepared.theta_values:
        if np.isclose(cur_theta, 0.0):
            img_rot = prepared.padded_image.copy()
        else:
            img_rot = img_as_ubyte(
                tran.rotate(prepared.padded_image, cur_theta, order=1, resize=False)
            )
            img_rot[img_rot >= 128] = 255
            img_rot[img_rot < 128] = 0

            if config.reskeletonize:
                filt_params = [2, 1, 1]
                img_rot = sdrv.apply_driver_morph(img_rot, filt_params, quiet_in=True)
                img_bool = img_as_bool(img_rot)
                img_rot = img_as_ubyte(morph.skeletonize(img_bool))

        theta_distances: List[float] = []
        theta_inverse: List[float] = []

        for rr in range(0, prepared.padded_image.shape[0], prepared.row_step):
            cur_row = img_rot[rr]
            segment_indices = gsz.find_intersections(cur_row)
            for seg_start, seg_end in segment_indices:
                pix_dist = seg_end - seg_start
                micron_dist = pix_dist / prepared.pixels_per_micron
                if micron_dist <= 0:
                    continue
                theta_distances.append(micron_dist)
                theta_inverse.append(1.0 / micron_dist)
                distances.append((float(cur_theta), micron_dist))
                inverse_distances.append((float(cur_theta), 1.0 / micron_dist))
                img_rot[rr, seg_start:seg_end] = 150

        per_theta_distances.append(np.array(theta_distances))
        per_theta_inverse.append(np.array(theta_inverse))
        rotated_images.append(img_rot.copy())

    distances_arr = (
        np.array(distances) if distances else np.empty((0, 2), dtype=float)
    )
    inverse_distances_arr = (
        np.array(inverse_distances) if inverse_distances else np.empty((0, 2), dtype=float)
    )

    theta_labels = [f"{theta:.1f}" for theta in prepared.theta_values]

    return LineMeasurementResult(
        distances=distances_arr,
        inverse_distances=inverse_distances_arr,
        per_theta_distances=per_theta_distances,
        per_theta_inverse_distances=per_theta_inverse,
        theta_labels=theta_labels,
        rotated_images=rotated_images,
    )


def aggregate_statistics(
    measurements: LineMeasurementResult, config: LineGridConfig
) -> StatisticsResult:
    """Compute statistics for each rotation and aggregate across angles.

    Parameters
    ----------
    measurements : LineMeasurementResult
        Measured intersection distances produced by :func:`measure_line_intersections`.
    config : LineGridConfig
        Pipeline configuration providing metadata for the summary row.

    Returns
    -------
    StatisticsResult
        Per-angle statistics, combined statistics and tabular artefacts.

    Raises
    ------
    ValueError
        Propagated if invalid numeric operations occur during aggregation.

    Examples
    --------
    >>> aggregate_statistics(measurements, config)  # doctest: +SKIP
    StatisticsResult(...)
    """

    properties = [
        "Total Number of Grain Segments",
        "Summed Length of Grain Segments (µm)",
        "Average Grain Size (µm)",
        "ASTM G (from l̄)",
        "Median Grain Size (µm)",
        "Std. Deviation in Grain Size (µm)",
        "Thickness (Average inverse Grain Size) (µm)",
        "Thickness (Median inverse Grain Size) (µm)",
        "95% CI half-width (µm)",
        "95% CI low (µm)",
        "95% CI high (µm)",
        "Relative Accuracy % (95% CI)",
    ]

    results_dict: Dict[str, List[object]] = {"Property": properties}
    angle_stats: List[AngleStatistics] = []

    for label, dist_arr, inv_arr in zip(
        measurements.theta_labels,
        measurements.per_theta_distances,
        measurements.per_theta_inverse_distances,
    ):
        stats = _compute_angle_statistics(label, dist_arr, inv_arr)
        angle_stats.append(stats)
        results_dict[label] = [
            stats.segment_count,
            stats.total_length,
            stats.average_length,
            stats.astm_g,
            stats.median_length,
            stats.std_dev,
            stats.thickness_from_average,
            stats.thickness_from_median,
            stats.ci95_halfwidth_um,
            stats.ci95_low_um,
            stats.ci95_high_um,
            stats.rel_accuracy_pct,
        ]

    non_empty_distances = [arr for arr in measurements.per_theta_distances if arr.size]
    non_empty_inverse = [arr for arr in measurements.per_theta_inverse_distances if arr.size]
    all_distances = (
        np.concatenate(non_empty_distances) if non_empty_distances else np.array([])
    )
    all_inverse = (
        np.concatenate(non_empty_inverse) if non_empty_inverse else np.array([])
    )
    overall_stats = _compute_angle_statistics("All Lines", all_distances, all_inverse)
    results_dict["All Lines"] = [
        overall_stats.segment_count,
        overall_stats.total_length,
        overall_stats.average_length,
        overall_stats.astm_g,
        overall_stats.median_length,
        overall_stats.std_dev,
        overall_stats.thickness_from_average,
        overall_stats.thickness_from_median,
        overall_stats.ci95_halfwidth_um,
        overall_stats.ci95_low_um,
        overall_stats.ci95_high_um,
        overall_stats.rel_accuracy_pct,
    ]

    results_table = pd.DataFrame(results_dict)
    distances_df = pd.DataFrame(all_distances, columns=["Distances (µm)"])
    inverse_distances_df = pd.DataFrame(
        all_inverse, columns=["Inverse Distances (1/µm)"]
    )

    summary_row = pd.DataFrame(
        [
            {
                "Filename": config.file_in_path.name,
                "Avg Grain Size (µm)": overall_stats.average_length,
                "ASTM G (from l̄)": overall_stats.astm_g,
                "Med. Grain Size (µm)": overall_stats.median_length,
                "Thickness (Avg inv. gs) (µm)": overall_stats.thickness_from_average,
                "Thickness (Med. inv. gs) (µm)": overall_stats.thickness_from_median,
                "Inverse Int Avg": overall_stats.average_inverse,
                "Inverse Int Med": overall_stats.median_inverse,
                "95% CI half-width (µm)": overall_stats.ci95_halfwidth_um,
                "%RA (95% CI)": overall_stats.rel_accuracy_pct,
            }
        ]
    )

    return StatisticsResult(
        angle_statistics=angle_stats,
        overall_statistics=overall_stats,
        results_table=results_table,
        distances_df=distances_df,
        inverse_distances_df=inverse_distances_df,
        summary_row=summary_row,
    )


def save_outputs(
    prepared: PreparedImageData,
    measurements: LineMeasurementResult,
    statistics: StatisticsResult,
    config: LineGridConfig,
    options: Optional[SaveOptions] = None,
) -> SaveArtifacts:
    """Persist selected artefacts produced by the pipeline.

    Parameters
    ----------
    prepared : PreparedImageData
        Pre-processed images and output paths.
    measurements : LineMeasurementResult
        Distances and rotated images produced by :func:`measure_line_intersections`.
    statistics : StatisticsResult
        Aggregated statistics produced by :func:`aggregate_statistics`.
    config : LineGridConfig
        Pipeline configuration describing file names and metadata.
    options : SaveOptions, optional
        Flags controlling which artefacts are persisted. ``None`` creates defaults.

    Returns
    -------
    SaveArtifacts
        Paths to the artefacts that were written to disk.

    Raises
    ------
    OSError
        Propagated when writing artefacts fails.

    Examples
    --------
    >>> save_outputs(prepared, measurements, statistics, config)  # doctest: +SKIP
    SaveArtifacts(...)
    """

    options = options or SaveOptions()
    artifacts = SaveArtifacts()

    prepared.results_dir.mkdir(parents=True, exist_ok=True)
    if options.save_rotated_images:
        prepared.rotated_dir.mkdir(parents=True, exist_ok=True)
        artifacts.rotated_image_paths = _save_rotated_images(
            measurements.rotated_images,
            measurements.theta_labels,
            prepared.rotated_dir,
            prepared.base_output_path.stem,
        )

    if options.save_boxplot:
        artifacts.boxplot_path = _save_boxplot(
            measurements.per_theta_distances,
            measurements.theta_labels,
            prepared.base_output_path.with_name(prepared.base_output_path.name + "_boxplot"),
        )

    if options.save_histograms:
        artifacts.histogram_paths = _save_histograms(
            statistics, prepared.base_output_path
        )

    if options.save_excel:
        artifacts.excel_path = _save_excel(
            statistics,
            prepared.base_output_path.with_suffix(".xlsx"),
            config.to_input_parameters(),
        )

    if options.append_summary:
        artifacts.summary_excel_path = _append_summary_excel(
            statistics.summary_row, config.summary_excel_path
        )

    if options.show_plots:
        plt.show()
    else:
        plt.close("all")

    return artifacts


def process_image(
    config: LineGridConfig, options: Optional[SaveOptions] = None
) -> Tuple[StatisticsResult, SaveArtifacts]:
    """Execute the full measurement pipeline and persist artefacts.

    Parameters
    ----------
    config : LineGridConfig
        Configuration describing the input image and processing parameters.
    options : SaveOptions, optional
        Flags controlling which artefacts are stored.

    Returns
    -------
    tuple of (:class:`StatisticsResult`, :class:`SaveArtifacts`)
        Final statistics and references to written artefacts.

    Raises
    ------
    FileNotFoundError
        If the input image does not exist.
    OSError
        Propagated when saving artefacts fails.

    Examples
    --------
    >>> process_image(config)  # doctest: +SKIP
    (StatisticsResult(...), SaveArtifacts(...))
    """

    print("\nCounting intersect distances for each orientation...")
    prepared = prepare_image(config)
    measurements = measure_line_intersections(prepared, config)
    describe_measurements(measurements)
    statistics = aggregate_statistics(measurements, config)
    print_statistics(statistics)
    artifacts = save_outputs(prepared, measurements, statistics, config, options)
    print("\nScript successfully terminated!")
    return statistics, artifacts


def _compute_angle_statistics(
    angle_label: str, distances: np.ndarray, inverse_distances: np.ndarray
) -> AngleStatistics:
    if distances.size == 0:
        return AngleStatistics(
            angle_label=angle_label,
            segment_count=0,
            total_length=0.0,
            average_length=float("nan"),
            median_length=float("nan"),
            std_dev=float("nan"),
            average_inverse=float("nan"),
            median_inverse=float("nan"),
            thickness_from_average=float("nan"),
            thickness_from_median=float("nan"),
            astm_g=float("nan"),
            ci95_halfwidth_um=float("nan"),
            ci95_low_um=float("nan"),
            ci95_high_um=float("nan"),
            rel_accuracy_pct=float("nan"),
        )

    total_length = float(np.sum(distances))
    average_length = float(np.mean(distances))
    median_length = float(np.median(distances))
    std_dev = float(np.std(distances))
    average_inverse = float(np.mean(inverse_distances)) if inverse_distances.size else float("nan")
    median_inverse = float(np.median(inverse_distances)) if inverse_distances.size else float("nan")

    thickness_avg = float(1.0 / (1.5 * average_inverse)) if average_inverse > 0 else float("nan")
    thickness_med = float(1.0 / (1.5 * median_inverse)) if median_inverse > 0 else float("nan")
    astm_g = astm_g_from_lbar_um(average_length)

    ci95_halfwidth = float("nan")
    ci95_low = float("nan")
    ci95_high = float("nan")
    rel_accuracy = float("nan")
    n = int(distances.size)
    if n >= 2 and np.isfinite(average_length):
        sample_std = float(np.std(distances, ddof=1))
        df = n - 1
        t_multiplier = t95_from_df(df)
        ci95_halfwidth = t_multiplier * sample_std / np.sqrt(n)
        ci95_low = average_length - ci95_halfwidth
        ci95_high = average_length + ci95_halfwidth
        if average_length > 0:
            rel_accuracy = 100.0 * ci95_halfwidth / average_length

    return AngleStatistics(
        angle_label=angle_label,
        segment_count=int(distances.size),
        total_length=total_length,
        average_length=average_length,
        median_length=median_length,
        std_dev=std_dev,
        average_inverse=average_inverse,
        median_inverse=median_inverse,
        thickness_from_average=thickness_avg,
        thickness_from_median=thickness_med,
        astm_g=astm_g,
        ci95_halfwidth_um=ci95_halfwidth,
        ci95_low_um=ci95_low,
        ci95_high_um=ci95_high,
        rel_accuracy_pct=rel_accuracy,
    )


def _save_rotated_images(
    rotated_images: Sequence[np.ndarray],
    theta_labels: Sequence[str],
    output_dir: Path,
    file_stem: str,
) -> List[Path]:
    paths: List[Path] = []
    for img, label in zip(rotated_images, theta_labels):
        file_out = output_dir / f"{file_stem}_{int(round(float(label)))}.tif"
        save_flag = imex.save_image(img, str(file_out))
        if save_flag:
            paths.append(file_out)
    return paths


def _save_boxplot(
    per_theta_distances: Sequence[np.ndarray],
    theta_labels: Sequence[str],
    output_path: Path,
) -> Optional[Path]:
    valid_data = [
        (label, data)
        for label, data in zip(theta_labels, per_theta_distances)
        if data.size
    ]
    if not valid_data:
        return None

    labels, data = zip(*valid_data)
    fig, ax = plt.subplots()
    ax.boxplot(data, sym="", labels=list(labels))

    for m, cur_y_arr in enumerate(data):
        cur_x_arr = np.random.normal(m + 1, 0.05, size=cur_y_arr.size)
        ax.plot(cur_x_arr, cur_y_arr, ".b", alpha=0.2)

    ax.set_xlabel("Rotation of Line Intercepts [degrees]")
    ax.set_ylabel("Segment Length of Grains [µm]")

    output_path = output_path.with_suffix(".png")
    fig.savefig(output_path)
    plt.close(fig)
    return output_path


def _save_histograms(
    statistics: StatisticsResult, base_output_path: Path
) -> List[Path]:
    paths: List[Path] = []
    if not statistics.distances_df.empty:
        path = base_output_path.with_name(base_output_path.name + "_histogram_distances").with_suffix(".png")
        _create_histogram(
            statistics.distances_df["Distances (µm)"].to_numpy(),
            "Histogram of Distances",
            "Distance (µm)",
            path,
        )
        paths.append(path)

    if not statistics.inverse_distances_df.empty:
        path = base_output_path.with_name(base_output_path.name + "_histogram_inverse_distances").with_suffix(".png")
        _create_histogram(
            statistics.inverse_distances_df["Inverse Distances (1/µm)"].to_numpy(),
            "Histogram of Inverse Distances",
            "Inverse Distance (1/µm)",
            path,
        )
        paths.append(path)

    return paths


def _create_histogram(data: np.ndarray, title: str, xlabel: str, output_path: Path) -> None:
    fig, ax = plt.subplots()
    ax.hist(data, bins=30, edgecolor="black")
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel("Frequency")
    ax.grid(True)
    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)


def _save_excel(
    statistics: StatisticsResult, output_path: Path, input_parameters: Dict[str, object]
) -> Path:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        statistics.results_table.to_excel(writer, sheet_name="Results", index=False)
        statistics.distances_df.to_excel(writer, sheet_name="Distances", index=False)
        statistics.inverse_distances_df.to_excel(
            writer, sheet_name="Inverse Distances", index=False
        )

    wb = load_workbook(output_path)
    ws = wb["Results"]

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.number_format = "0.00"

    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="Input Parameters")
    for i, (param, value) in enumerate(input_parameters.items(), start=start_row + 1):
        if isinstance(value, (tuple, list)):
            value = str(value)
        ws.cell(row=i, column=1, value=param)
        ws.cell(row=i, column=2, value=value)

    wb.save(output_path)
    return output_path


def _append_summary_excel(summary_row: pd.DataFrame, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not output_path.exists():
        summary_row.to_excel(output_path, index=False, engine="openpyxl")
    else:
        with pd.ExcelWriter(
            output_path, mode="a", engine="openpyxl", if_sheet_exists="overlay"
        ) as writer:
            sheet = writer.sheets.get("Sheet1")
            start_row = sheet.max_row if sheet is not None else 0
            summary_row.to_excel(
                writer, index=False, header=False, startrow=start_row
            )

    wb = load_workbook(output_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.number_format = "0.00"
    wb.save(output_path)
    return output_path

